
from openpyxl import Workbook, workbook
from openpyxl import load_workbook
import openpyxl
import requests
import pandas as pd
import datetime
import os

# print(openpyxl.__version__)
L=[1,3,5,10,20]

def append_df_to_excel(filename, df, sheet_name='Sheet1', startrow=None,
                       truncate_sheet=False,**to_excel_kwargs):
    """
    Append a DataFrame [df] to existing Excel file [filename]
    into [sheet_name] Sheet.
    If [filename] doesn't exist, then this function will create it.
    Parameters:
      filename : File path or existing ExcelWriter
                 (Example: '/path/to/file.xlsx')
      df : dataframe to save to workbook
      sheet_name : Name of sheet which will contain DataFrame.
                   (default: 'Sheet1')
      startrow : upper left cell row to dump data frame.
                 Per default (startrow=None) calculate the last row
                 in the existing DF and write to the next row...
      truncate_sheet : truncate (remove and recreate) [sheet_name]
                       before writing DataFrame to Excel file
      to_excel_kwargs : arguments which will be passed to `DataFrame.to_excel()`
                        [can be dictionary]
    Returns: None
    """
    if 'engine' in to_excel_kwargs:
        to_excel_kwargs.pop('engine')

    writer = pd.ExcelWriter(filename, engine='openpyxl')
    try:
        FileNotFoundError1
    except:
        # FileNotFoundError = IOError
        pass

    # if os.path.exists(filename):
    #     new_wb=load_workbook(filename)
    # else:
    #     new_wb=Workbook()

    # try:
        # try to open an existing workbook
        wb = openpyxl.load_workbook(filename)
        # writer.book = load_workbook(filename)
        print(wb.sheetnames)

    #     # get the last row in the existing Excel sheet
    #     # if it was not specified explicitly
    #     if startrow is None and sheet_name in writer.book.sheetnames:
    #         startrow = writer.book[sheet_name].max_row

    #     # truncate sheet
    #     if truncate_sheet and sheet_name in writer.book.sheetnames:
    #         # index of [sheet_name] sheet
    #         idx = writer.book.sheetnames.index(sheet_name)
    #         # remove [sheet_name]
    #         writer.book.remove(writer.book.worksheets[idx])
    #         # create an empty sheet [sheet_name] using old index
    #         writer.book.create_sheet(sheet_name, idx)

    #     # copy existing sheets
    #     writer.sheets = {ws.title: ws for ws in writer.book.worksheets}
    # except FileNotFoundError:
    #     # file does not exist yet, we will create it
    #     pass

    # if startrow is None:
    #     startrow = 0

    # # write out the new sheet
    # df.to_excel(writer, sheet_name, startrow=startrow, **to_excel_kwargs)
    # # save the workbook
    # writer.save()
    # writer.close()

j=0
for i in L:
    res1=requests.get(f"https://fubon-ebrokerdj.fbs.com.tw/z/zg/zg_F_0_{str(i)}.djhtm")
    res2=requests.get(f"https://fubon-ebrokerdj.fbs.com.tw/z/zg/zg_DD_0_{str(i)}.djhtm")
    res3=requests.get(f"https://fubon-ebrokerdj.fbs.com.tw/z/zg/zg_F_0_{str(i)}.djhtm")
    res4=requests.get(f"https://fubon-ebrokerdj.fbs.com.tw/z/zg/zg_D_1_{str(i)}.djhtm")
    res5=requests.get(f"https://fubon-ebrokerdj.fbs.com.tw/z/zg/zg_DD_1_{str(i)}.djhtm")
    res6=requests.get(f"https://fubon-ebrokerdj.fbs.com.tw/z/zg/zg_F_1_{str(i)}.djhtm")

    dfs1=pd.read_html(res1.text)
    dfs2=pd.read_html(res2.text)
    dfs3=pd.read_html(res3.text)
    df1=dfs1[2][1]    
    df2=dfs2[2][1]
    df3=dfs3[2][1]
    df4=pd.merge(df1,df2)
    df5=pd.merge(df4,df3)

    dfs6=pd.read_html(res4.text)
    dfs7=pd.read_html(res5.text)
    dfs8=pd.read_html(res6.text)
    df6=dfs6[2][1]    
    df7=dfs7[2][1]
    df8=dfs8[2][1]
    df9=pd.merge(df6,df7)
    df10=pd.merge(df8,df9)

    filename=fr"C:\Users\AI\Desktop\2022\跟蹤主力_{str(datetime.date.today())}.xlsx"
    append_df_to_excel(filename,df4,sheet_name='外投主合買', startcol=j,startrow=0,index=False)
    append_df_to_excel(filename,df5,sheet_name='4是和一', startcol=j,startrow=0,index=False)
    append_df_to_excel(filename,df9,sheet_name='3貴和一', startcol=j,startrow=0,index=False)
    append_df_to_excel(filename,df10,sheet_name='3貴和一_主', startcol=j,startrow=0,index=False)

    j+=1


    print("***************************************************************")
    print(i,'4是和一    ===> 合 ',df5)
    print("***************************************************************")
    print(i,'3是和一    ===> 合 ',df4)
    print("***************************************************************")

    print("***************************************************************")
    print(i,'4貴和一    ===> 合 ',df10)
    print("***************************************************************")
    print(i,'3貴和一    ===> 合 ',df9)
    print("***************************************************************")

